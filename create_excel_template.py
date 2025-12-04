"""
Create an Excel template with data validation (dropdown lists) for bulk upload
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

def create_excel_template():
    """Create Excel template with data validation"""
    
    # Define options for dropdowns
    type_options = ['Customer', 'Technician']
    
    service_type_options = [
        'Personal care',
        'Domestic Assistance',
        'Community Access',
        'Transport',
        'Behaviour Support',
        'Support Coordination',
        'Therapy Access',
        'Assistive Tech',
        'Life Skills Training'
    ]
    
    skill_options = [
        'Personal care',
        'Domestic Assistance',
        'Community Access',
        'Transport',
        'Behaviour Support',
        'Support Coordination',
        'Therapy Access',
        'Assistive Tech',
        'Life Skills Training'
    ]
    
    priority_options = ['High', 'Medium', 'Low']
    
    color_options = ['#4285F4', '#34A853', '#FBBC05', '#EA4335', '#FF6D01']
    
    # Create sample data
    data = {
        'Type': ['Customer', 'Technician'],
        'Username': ['', ''],
        'Email': ['', ''],
        'Password': ['Welcome123', 'Welcome123'],
        'Phone': ['', ''],
        'Address': ['', ''],
        'ServiceMinutes': [60, ''],
        'WindowStart': ['', ''],
        'WindowEnd': ['', ''],
        'RequiredSkills': ['', ''],
        'Priority': ['Medium', ''],
        'ServiceType': ['', ''],
        'DepotAddress': ['', ''],
        'CapacityHours': ['', 8],
        'ShiftStart': ['', '08:00'],
        'ShiftEnd': ['', '17:00'],
        'Skills': ['', ''],
        'ColorHex': ['', '#4285F4']
    }
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Save to Excel
    output_file = 'bulk_upload_template_with_dropdowns.xlsx'
    df.to_excel(output_file, index=False, sheet_name='Bulk Upload')
    
    # Load workbook for formatting
    wb = load_workbook(output_file)
    ws = wb['Bulk Upload']
    
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Auto-adjust column widths
        ws.column_dimensions[cell.column_letter].width = 15
    
    # Add data validation for each column
    print("Adding data validation...")
    
    # Type column (A, starting from row 2)
    type_dv = DataValidation(type="list", formula1=f'"{",".join(type_options)}"')
    type_dv.add(f'A2:A{ws.max_row + 50}')  # Allow up to 50 rows
    ws.add_data_validation(type_dv)
    
    # ServiceType column (L, starting from row 2)
    service_type_dv = DataValidation(type="list", formula1=f'"{",".join(service_type_options)}"')
    service_type_dv.add(f'L2:L{ws.max_row + 50}')
    ws.add_data_validation(service_type_dv)
    
    # RequiredSkills column (J, starting from row 2)
    skills_dv = DataValidation(type="list", formula1=f'"{",".join(skill_options)}"')
    skills_dv.add(f'J2:J{ws.max_row + 50}')
    ws.add_data_validation(skills_dv)
    
    # Skills column (for technicians) (Q, starting from row 2)
    tech_skills_dv = DataValidation(type="list", formula1=f'"{",".join(skill_options)}"')
    tech_skills_dv.add(f'Q2:Q{ws.max_row + 50}')
    ws.add_data_validation(tech_skills_dv)
    
    # Priority column (K, starting from row 2)
    priority_dv = DataValidation(type="list", formula1=f'"{",".join(priority_options)}"')
    priority_dv.add(f'K2:K{ws.max_row + 50}')
    ws.add_data_validation(priority_dv)
    
    # ColorHex column (R, starting from row 2)
    color_dv = DataValidation(type="list", formula1=f'"{",".join(color_options)}"')
    color_dv.add(f'R2:R{ws.max_row + 50}')
    ws.add_data_validation(color_dv)
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions", 0)
    
    instructions = [
        ["Bulk Upload Excel Template - Instructions"],
        [""],
        ["OVERVIEW", ""],
        ["This template allows you to bulk upload Customer and Technician data to the system.", ""],
        ["", ""],
        ["REQUIRED FIELDS", ""],
        ["Common (all): Type, Username, Email", ""],
        ["Customer: Address, ServiceType, WindowStart, WindowEnd", ""],
        ["Technician: DepotAddress, ShiftStart, ShiftEnd", ""],
        ["", ""],
        ["FIELD DESCRIPTIONS", ""],
        ["Type: Customer or Technician (use dropdown)", ""],
        ["Username: Unique username for the user", ""],
        ["Email: Valid email address", ""],
        ["Password: Default is Welcome123 if not specified", ""],
        ["Phone: Optional phone number", ""],
        ["Address: Melbourne address for service location (Customer only)", ""],
        ["ServiceMinutes: Duration in minutes (15-480)", ""],
        ["WindowStart: When service can start (YYYY-MM-DD HH:MM)", ""],
        ["WindowEnd: When service must complete (YYYY-MM-DD HH:MM)", ""],
        ["RequiredSkills: Choose from dropdown (Customer only)", ""],
        ["Priority: High, Medium, or Low (use dropdown)", ""],
        ["ServiceType: Choose from dropdown (Customer only)", ""],
        ["DepotAddress: Melbourne depot address (Technician only)", ""],
        ["CapacityHours: Daily work hours 1-24 (Technician only)", ""],
        ["ShiftStart: Work start time HH:MM (Technician only)", ""],
        ["ShiftEnd: Work end time HH:MM (Technician only)", ""],
        ["Skills: Comma-separated skills (Technician only, use dropdown)", ""],
        ["ColorHex: Hex color for route display (use dropdown)", ""],
        ["", ""],
        ["EXAMPLES", ""],
        ["Customer Example:", ""],
        ["Type=Customer, Username=john_doe, Email=john@example.com", ""],
        ["Address=123 Collins St, Melbourne VIC 3000", ""],
        ["WindowStart=2025-11-01 09:00, WindowEnd=2025-11-01 17:00", ""],
        ["RequiredSkills=Personal care, ServiceType=Personal care", ""],
        ["", ""],
        ["Technician Example:", ""],
        ["Type=Technician, Username=tech_john, Email=tech@example.com", ""],
        ["DepotAddress=456 Bourke St, Melbourne VIC 3000", ""],
        ["ShiftStart=08:00, ShiftEnd=17:00, CapacityHours=8", ""],
        ["Skills=Personal care,Transport, ColorHex=#4285F4", ""],
        ["", ""],
        ["NOTES", ""],
        ["- All addresses must be in Melbourne, VIC, Australia", ""],
        ["- Skills must match the dropdown options exactly", ""],
        ["- Dates should be in format: YYYY-MM-DD HH:MM", ""],
        ["- Delete the example rows before uploading your data", ""],
        ["- Ensure no blank rows between data entries", ""],
    ]
    
    for row in instructions:
        ws_instructions.append(row)
    
    # Format instructions sheet
    ws_instructions.column_dimensions['A'].width = 60
    ws_instructions.column_dimensions['B'].width = 80
    ws_instructions['A1'].font = Font(size=14, bold=True)
    
    # Add helper formulas sheet
    ws_formulas = wb.create_sheet("Helper Formulas", 1)
    
    formulas_data = [
        ["Helper Formulas", ""],
        ["", ""],
        ["SKILLS LIST (Copy to use as reference)", ""],
    ]
    formulas_data.extend([[skill, ""] for skill in skill_options])
    
    formulas_data.extend([
        [""],
        ["SERVICE TYPES (Copy to use as reference)", ""],
    ])
    formulas_data.extend([[stype, ""] for stype in service_type_options])
    
    for row in formulas_data:
        ws_formulas.append(row)
    
    ws_formulas.column_dimensions['A'].width = 30
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ Excel template created: {output_file}")
    print(f"✓ Contains: Dropdown lists, Instructions, Helper data")
    print(f"✓ Ready to use for bulk upload!")

if __name__ == '__main__':
    create_excel_template()

